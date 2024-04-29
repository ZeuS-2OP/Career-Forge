import openpyxl
from collections import namedtuple
from math import isclose
import csv
# Define a named tuple to store student data
Student = namedtuple('Student', ['name', 'academic_grades', 'advanced_courses', 'research_publication',
                                 'awards_scholarship', 'coding_competition', 'contribution_projects', 'internships'])

def extract_students_from_excel(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return []

    worksheet = workbook.active

    # Extract headers
    headers = []
    for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)):
        if isinstance(cell, openpyxl.cell.cell.Cell):
            headers.append(cell.value)
        else:
            headers.append(str(cell))

    students = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        student_data = dict(zip(headers, row))
        student = Student(
            name=student_data['name'],
            academic_grades=int(student_data['academic_grades']),
            advanced_courses=float(student_data['advanced_courses']),
            research_publication=int(student_data['research_publication']),
            awards_scholarship=int(student_data['awards_scholarship']),
            coding_competition=int(student_data['coding_competition']),
            contribution_projects=int(student_data['contribution_projects']),
            internships=int(student_data['internships'])
        )
        students.append(student)

    return students

def build_decision_tree(criteria, weights):
    root = DecisionNode('academic_grades', weights)

    # Create the nodes for academic grades
    academic_grades_node = root
    academic_grades_ranges = [(4.0, 'excellent'), (3.5, 'good'), (3.0, 'average'), (0.0, 'poor')]
    for lower_bound, grade_label in academic_grades_ranges:
        child_node = DecisionNode('advanced_courses', weights)
        academic_grades_node.add_child(grade_label, child_node)
        academic_grades_node = child_node

    # Create the nodes for advanced courses
    advanced_courses_nodes = academic_grades_node.children.values()
    for advanced_courses_node in advanced_courses_nodes:
        advanced_courses_values = [3, 2, 1, 0]
        for value in advanced_courses_values:
            child_node = DecisionNode('research_publication', weights)
            advanced_courses_node.add_child(value, child_node)

    # Create the nodes for research and publication
    research_publication_nodes = [node for advanced_courses_node in advanced_courses_nodes for node in advanced_courses_node.children.values()]
    for research_publication_node in research_publication_nodes:
        research_publication_values = [4, 3, 2, 1, 0]
        for value in research_publication_values:
            child_node = DecisionNode('awards_scholarship', weights)
            research_publication_node.add_child(value, child_node)

    # Create the nodes for awards and scholarship
    awards_scholarship_nodes = [node for research_publication_node in research_publication_nodes for node in research_publication_node.children.values()]
    for awards_scholarship_node in awards_scholarship_nodes:
        awards_scholarship_values = [3, 2, 1, 0]
        for value in awards_scholarship_values:
            child_node = DecisionNode('coding_competition', weights)
            awards_scholarship_node.add_child(value, child_node)

    # Create the nodes for coding competition
    coding_competition_nodes = [node for awards_scholarship_node in awards_scholarship_nodes for node in awards_scholarship_node.children.values()]
    for coding_competition_node in coding_competition_nodes:
        coding_competition_values = [3, 2, 1, 0]
        for value in coding_competition_values:
            child_node = DecisionNode('contribution_projects', weights)
            coding_competition_node.add_child(value, child_node)

    # Create the nodes for contribution to projects
    contribution_projects_nodes = [node for coding_competition_node in coding_competition_nodes for node in coding_competition_node.children.values()]
    for contribution_projects_node in contribution_projects_nodes:
        contribution_projects_values = [2, 1, 0]
        for value in contribution_projects_values:
            child_node = DecisionNode('internships', weights)
            contribution_projects_node.add_child(value, child_node)

    # Create the leaf nodes for internships
    internships_nodes = [node for contribution_projects_node in contribution_projects_nodes for node in contribution_projects_node.children.values()]
    for internships_node in internships_nodes:
        internships_values = [2, 1, 0]
        for value in internships_values:
            leaf_node = DecisionNode(None, None, is_leaf=True)
            leaf_node.score = internships_node.compute_score(Student(None, 0, 0, 0, 0, 0, 0, value))
            internships_node.add_child(value, leaf_node)

    return root

class DecisionNode:
    def __init__(self, criterion, weights, is_leaf=False, score=None):
        self.criterion = criterion
        self.weights = weights
        self.is_leaf = is_leaf
        self.score = score
        self.children = {}

    def add_child(self, value, child):
        self.children[value] = child

    def evaluate(self, student_data):
        if self.is_leaf:
            return self.score
        else:
            student_value = getattr(student_data, self.criterion)
            if student_value in self.children:
                return self.children[student_value].evaluate(student_data)
            else:
                return self.compute_score(student_data)

    def compute_score(self, student_data):
        composite_score = 0
        for criterion, weight in self.weights.items():
            student_value = getattr(student_data, criterion)
            composite_score += student_value * weight
        return composite_score


def rank_students(students, decision_tree):
    student_scores = []
    for student in students:
        composite_score = decision_tree.evaluate(student)
        student_scores.append((student, composite_score))

    # Sort students by composite score in descending order
    ranked_students = sorted(student_scores, key=lambda x: x[1], reverse=True)

    # Assign ranks
    ranked_with_ranks = []
    rank = 1
    prev_score = 0
    for student, score in ranked_students:
        if not isclose(score, prev_score):
            prev_score = score
            rank = len(ranked_with_ranks) + 1
        ranked_with_ranks.append((student, rank))

    return ranked_with_ranks
    return student_scores

# Prompt the user for the file path

# Extract students from the Excel file
students = extract_students_from_excel('algosheetex1.xlsx')

if not students:
    print("No student data found.")
else:
    # Define criteria weights
    criteria_weights = {
        'academic_grades': 0.4,
        'advanced_courses': 0.15,
        'research_publication': 0.15,
        'awards_scholarship': 0.1,
        'coding_competition': 0.1,
        'contribution_projects': 0.05,
        'internships': 0.05
    }

    # Build the decision tree
    decision_tree = build_decision_tree(criteria_weights.keys(), criteria_weights)

    ranked_students = rank_students(students, decision_tree)

    print("\nRanked Students:")
    count  = 0
    for student, rank in ranked_students:
        print(f"Rank {count}: {student.name} (Academic Grades: {student.academic_grades}, Advanced Courses: {student.advanced_courses}, "
              f"Research & Publication: {student.research_publication}, Awards & Scholarship: {student.awards_scholarship}, "
              f"Coding Competition: {student.coding_competition}, Contribution to Projects: {student.contribution_projects}, "
              f"Internships: {student.internships})")
        count+=1

csv_file_path = "ranked_students.csv"

# Write the ranked students data to a CSV file
with open(csv_file_path, mode='w', newline='') as file:
    writer = csv.writer(file)
    writer.writerow(["Rank", "Name", "Academic Grades", "Advanced Courses", "Research & Publication",
                     "Awards & Scholarship", "Coding Competition", "Contribution to Projects", "Internships"])
    
    count = 1
    for student, rank in ranked_students:
        writer.writerow([count, student.name, student.academic_grades, student.advanced_courses,
                         student.research_publication, student.awards_scholarship, student.coding_competition,
                         student.contribution_projects, student.internships])
        count+=1
print(f"Ranked student data has been written to {csv_file_path}")